# -*- coding: utf-8 -*-
"""炸板阴线策略回测：T 日炸板阴线，T+1 竞价买入，T+2 收盘卖出。

这是用户确认过的正确交易规则（A 股 T+1 无法当天卖出）：
  T    筛选日：盘中触板但收盘未封板（炸板），且当日收阴线
  T+1  次日竞价买入（以开盘价成交）
  T+2  第三日收盘卖出

支持两类数据源：
  1. 聚宽导出 CSV（data/全市场A股_2025-04-21_2026-04-28.csv）
  2. baostock 导出 CSV（data/全市场A股_2026-07-01_2026-07-31_baostock.csv）

用法:
    python backtest/run_zhaban_t2.py
"""

from __future__ import annotations

import io
import os
import sys
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JQ_CSV = os.path.join(ROOT, "data", "全市场A股_2025-04-21_2026-04-28.csv")
BS_CSV = os.path.join(ROOT, "data", "全市场A股_2026-07-01_2026-07-31_baostock.csv")

# 单笔往返成本：佣金万2.5双边 + 滑点0.1%双边 + 卖出印花税0.05%
COST_PCT = 0.025 * 2 + 0.1 * 2 + 0.05


def _round_price(value: float, pct: float) -> float:
    """按交易所四舍五入规则计算涨停价。"""
    return float(
        (Decimal(str(value)) * Decimal(str(1 + pct))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    )


def limit_pct(code: str, is_st: int) -> float:
    """按板块和 ST 状态返回涨跌幅限制。"""
    if code.startswith("bj."):
        return 0.30
    if code.startswith("sz.3") or code.startswith("sh.688"):
        return 0.20
    if is_st == 1:
        return 0.05
    return 0.10


def load_baostock(path: str) -> pd.DataFrame:
    """加载 baostock 日线并计算涨停价、T-1 涨停、炸板标记等因子。"""
    df = pd.read_csv(path, encoding="utf-8-sig", dtype={"code": str})
    for col in ["open", "high", "low", "close", "preclose", "amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["open", "close", "high", "preclose"])

    df["成交额_亿"] = (df["amount"] / 1e8).round(2)
    df["high_limit"] = [
        _round_price(row.preclose, limit_pct(row.code, int(row.isST)))
        for row in df.itertuples(index=False)
    ]
    df["is_limit_up"] = (df["close"] >= df["high_limit"] * 0.99).astype(int)
    df = df.sort_values(["code", "date"]).reset_index(drop=True)
    df["prev_limit_up"] = df.groupby("code")["is_limit_up"].shift(1).fillna(0).astype(int)
    df["auction_pct"] = (df["open"] / df["preclose"] - 1) * 100
    df["is_zhaban"] = (
        (df["high"] >= df["high_limit"] * 0.999)
        & (df["close"] < df["high_limit"])
        & (df["成交额_亿"] > 0.5)
        & (df["tradestatus"] == 1)
    ).astype(int)
    df.rename(columns={"date": "time"}, inplace=True)
    df["time"] = pd.to_datetime(df["time"])
    return df


def load_jq(path: str) -> pd.DataFrame:
    """加载聚宽 CSV，沿用下载脚本里已算好的因子。"""
    df = pd.read_csv(path, encoding="utf-8-sig", dtype={"code": str})
    df["time"] = pd.to_datetime(df["time"])
    df["is_zhaban"] = (
        (df["is_zhaban"] == 1) & (df["成交额_亿"] > 0.5)
    ).astype(int)
    return df


def backtest_t2(df: pd.DataFrame) -> pd.DataFrame:
    """计算 T+1 竞价买入、T+2 收盘卖出，只保留阴线炸板样本。"""
    df = df.sort_values(["code", "time"]).reset_index(drop=True)
    df["buy_price"] = df.groupby("code")["open"].shift(-1)
    df["sell_price"] = df.groupby("code")["close"].shift(-2)

    signal = (df["is_zhaban"] == 1) & (df["close"] < df["open"])
    out = df[signal].copy()
    out = out.dropna(subset=["buy_price", "sell_price"])
    out["盈亏_点"] = (out["sell_price"] - out["buy_price"]).round(2)
    out["盈亏_pct"] = (out["sell_price"] / out["buy_price"] - 1) * 100
    return out


def detail_table(out: pd.DataFrame) -> pd.DataFrame:
    """整理成用户习惯的明细表。"""
    return pd.DataFrame(
        {
            "日期": out["time"].dt.strftime("%Y-%m-%d"),
            "代码": out["code"],
            "炸板日收盘": out["close"],
            "阴线": 1,
            "成交额(亿)": out["成交额_亿"].round(2),
            "竞价涨幅(%)": out["auction_pct"].round(2),
            "T-1涨停": out["prev_limit_up"],
            "次日买入价": out["buy_price"],
            "第三日卖出价": out["sell_price"],
            "盈亏(点)": out["盈亏_点"],
            "盈亏(%)": out["盈亏_pct"].round(2),
        }
    )


def export_excel(detail: pd.DataFrame, xlsx_path: str, csv_path: str) -> None:
    """导出带样式的 Excel 和纯 CSV 明细表。"""
    detail.to_csv(csv_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        detail.to_excel(writer, index=False, sheet_name="明细")
        ws = writer.sheets["明细"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="D9E2F3")
        win_fill = PatternFill("solid", fgColor="C6EFCE")
        loss_fill = PatternFill("solid", fgColor="FFC7CE")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
            for cell in row:
                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue
                cell.fill = win_fill if value > 0 else loss_fill
                cell.number_format = "0.00"
        for idx, width in enumerate(
            [11, 14, 10, 6, 10, 11, 8, 10, 12, 9, 9], start=1
        ):
            ws.column_dimensions[get_column_letter(idx)].width = width


def trade_stats(out: pd.DataFrame, label: str) -> None:
    """打印一笔一算的整体绩效。"""
    n = len(out)
    if n == 0:
        print(f"{label}: 无样本")
        return
    ret = out["盈亏_pct"]
    win = (ret > 0).mean() * 100
    avg = ret.mean()
    avg_win = ret[ret > 0].mean() if (ret > 0).any() else 0.0
    avg_loss = ret[ret <= 0].mean() if (ret <= 0).any() else 0.0
    print(
        f"{label}: {n}笔, 胜率{win:.1f}%, 平均{avg:+.2f}%, "
        f"平均盈利{avg_win:+.2f}%, 平均亏损{avg_loss:+.2f}%, "
        f"最大{ret.max():+.2f}% / {ret.min():+.2f}%"
    )


def monthly_stats(out: pd.DataFrame) -> pd.DataFrame:
    """按 T+2 卖出日所在月份做等权复利汇总。"""
    rows = []
    for month, grp in out.groupby(out["time"].dt.to_period("M")):
        daily = grp.groupby(grp["time"].dt.date)["盈亏_pct"].mean()
        gross = ((1 + daily / 100).prod() - 1) * 100
        # 每笔等权时，组合每天只承担一笔往返成本
        net_daily = daily - COST_PCT
        net = ((1 + net_daily / 100).prod() - 1) * 100
        rows.append(
            {
                "月份": str(month),
                "笔数": len(grp),
                "胜率(%)": round((grp["盈亏_pct"] > 0).mean() * 100, 1),
                "平均收益(%)": round(grp["盈亏_pct"].mean(), 2),
                "等权毛收益(%)": round(gross, 2),
                "等权净收益(%)": round(net, 2),
                "10万净收益(元)": round(100000 * net / 100),
            }
        )
    return pd.DataFrame(rows)


def run_one(
    df: pd.DataFrame, label: str, prefix: str
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """跑单一数据源并导出明细。"""
    out = backtest_t2(df)
    detail = detail_table(out)
    export_excel(
        detail,
        os.path.join(ROOT, "results", f"阴线炸板明细表_{prefix}.xlsx"),
        os.path.join(ROOT, "results", f"阴线炸板明细表_{prefix}.csv"),
    )
    print()
    print("=" * 66)
    print(f"  {label}（T+1 竞价买入，T+2 收盘卖出）")
    print("=" * 66)
    trade_stats(out, "阴线炸板")
    month = monthly_stats(out)
    print(month.to_string(index=False))
    return out, month


def main() -> None:
    jq = load_jq(JQ_CSV) if os.path.isfile(JQ_CSV) else None
    bs_df = load_baostock(BS_CSV) if os.path.isfile(BS_CSV) else None

    all_parts: list[pd.DataFrame] = []
    months: list[pd.DataFrame] = []
    if jq is not None:
        out, month = run_one(jq, "全年 2025-04-21 ~ 2026-04-28", "2025-04-21_2026-04-28")
        all_parts.append(out)
        months.append(month)
    if bs_df is not None:
        out, month = run_one(bs_df, "2026年7月", "2026-07-01_2026-07-31")
        all_parts.append(out)
        months.append(month)

    if all_parts:
        combined = pd.concat(all_parts, ignore_index=True)
        print()
        trade_stats(combined, "合并样本")

    if months:
        summary = pd.concat(months, ignore_index=True)
        summary.to_csv(
            os.path.join(ROOT, "results", "炸板T2_月度汇总.csv"),
            index=False,
            encoding="utf-8-sig",
        )
        print(f"\n月度汇总已保存: results/炸板T2_月度汇总.csv")


if __name__ == "__main__":
    main()
