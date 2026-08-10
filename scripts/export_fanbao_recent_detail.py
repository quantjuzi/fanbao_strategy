# -*- coding: utf-8 -*-
"""导出近三个月反包模型实体涨幅明细表(Excel + CSV)。

数据源: baostock 日线(2026-04-01 ~ 2026-07-31)
反包信号: T-1 涨停(prev_limit_up == 1), T 日未涨停(is_limit_up == 0)
买入: T+1 开盘价; 卖出: T+2 收盘价
输出: results/反包实体明细表_2026-05-01_2026-07-31.xlsx / .csv
"""

import os
import sys
import io
from decimal import Decimal, ROUND_HALF_UP

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BS_FILES = [
    os.path.join(ROOT, "data", "全市场A股_2026-04-01_2026-06-30_baostock.csv"),
    os.path.join(ROOT, "data", "全市场A股_2026-07-01_2026-07-31_baostock.csv"),
]
OUT_XLSX = os.path.join(ROOT, "results", "反包实体明细表_2026-05-01_2026-07-31.xlsx")
OUT_CSV = os.path.join(ROOT, "results", "反包实体明细表_2026-05-01_2026-07-31.csv")


def _round_price(value: float, pct: float) -> float:
    """按交易所四舍五入规则计算涨停价。"""
    return float(
        (Decimal(str(value)) * Decimal(str(1 + pct))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    )


def _limit_pct(code: str, is_st: int) -> float:
    """按板块和 ST 状态返回涨跌幅限制。"""
    if code.startswith("bj."):
        return 0.30
    if code.startswith("sz.3") or code.startswith("sh.688"):
        return 0.20
    if is_st == 1:
        return 0.05
    return 0.10


def load_baostock(path: str) -> pd.DataFrame:
    """加载 baostock 日线并计算涨停价、T-1 涨停等因子。"""
    df = pd.read_csv(path, encoding="utf-8-sig", dtype={"code": str})
    for col in ["open", "high", "low", "close", "preclose", "amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["open", "close", "high", "preclose"])

    df["成交额_亿"] = (df["amount"] / 1e8).round(2)
    df["high_limit"] = [
        _round_price(row.preclose, _limit_pct(row.code, int(row.isST)))
        for row in df.itertuples(index=False)
    ]
    df["is_limit_up"] = (df["close"] >= df["high_limit"] * 0.99).astype(int)
    df = df.sort_values(["code", "date"]).reset_index(drop=True)
    df["prev_limit_up"] = (
        df.groupby("code")["is_limit_up"].shift(1).fillna(0).astype(int)
    )
    df["auction_pct"] = (df["open"] / df["preclose"] - 1) * 100
    df["time"] = pd.to_datetime(df["date"])
    return df


def main() -> None:
    """合并 5~7 月数据, 计算反包信号与实体/位置因子, 导出明细表。"""
    parts = [load_baostock(p) for p in BS_FILES if os.path.isfile(p)]
    df = pd.concat(parts, ignore_index=True)
    df = df.sort_values(["code", "time"]).reset_index(drop=True)

    # 近三个月: 2026-05-01 之后出现信号
    df = df[df["time"] >= "2026-05-01"].copy()
    df = df.sort_values(["code", "time"]).reset_index(drop=True)

    df["实体涨幅"] = (df["close"] / df["open"] - 1) * 100
    grp = df.groupby("code", sort=False)
    df["low20"] = grp["low"].transform(lambda s: s.rolling(20, min_periods=20).min())
    df["high20"] = grp["high"].transform(lambda s: s.rolling(20, min_periods=20).max())
    df["位置"] = (df["close"] - df["low20"]) / (df["high20"] - df["low20"])

    df["next_open"] = grp["open"].shift(-1)
    df["next2_close"] = grp["close"].shift(-2)

    signal = (df["prev_limit_up"] == 1) & (df["is_limit_up"] == 0)
    sig = df[signal].dropna(
        subset=["next_open", "next2_close", "low20", "high20"]
    ).copy()
    sig["盈亏点"] = sig["next2_close"] - sig["next_open"]
    sig["盈亏%"] = (sig["next2_close"] / sig["next_open"] - 1) * 100

    out = pd.DataFrame(
        {
            "日期": sig["time"].dt.strftime("%Y-%m-%d"),
            "代码": sig["code"],
            "T日收盘": sig["close"].round(2),
            "实体涨幅(%)": sig["实体涨幅"].round(2),
            "位置": sig["位置"].round(2),
            "成交额(亿)": sig["成交额_亿"].round(2),
            "竞价涨幅(%)": sig["auction_pct"].round(2),
            "次日买入价": sig["next_open"].round(2),
            "第三日卖出价": sig["next2_close"].round(2),
            "盈亏(点)": sig["盈亏点"].round(2),
            "盈亏(%)": sig["盈亏%"].round(2),
        }
    )
    out = out.sort_values(["日期", "代码"]).reset_index(drop=True)

    bins = [-np.inf, -5, -3, -1, 0, 3, 5, np.inf]
    labels = ["<-5% 大阴", "-5~-3", "-3~-1", "-1~0", "0~3 小阳", "3~5", ">5% 大红"]
    out["实体档位"] = pd.cut(out["实体涨幅(%)"], bins=bins, labels=labels, right=False)
    summary = (
        out.groupby("实体档位", observed=True)
        .agg(笔数=("盈亏(%)", "size"),
             胜率=("盈亏(%)", lambda x: round((x > 0).mean() * 100, 1)),
             平均收益=("盈亏(%)", lambda x: round(x.mean(), 2)),
             中位数=("盈亏(%)", lambda x: round(x.median(), 2)))
        .reset_index()
    )

    out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="明细")
        summary.to_excel(writer, index=False, sheet_name="实体汇总")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            header_fill = PatternFill("solid", fgColor="D9E2F3")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws = writer.sheets["明细"]
        win_fill = PatternFill("solid", fgColor="C6EFCE")
        loss_fill = PatternFill("solid", fgColor="FFC7CE")
        for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
            for cell in row:
                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue
                if value > 0:
                    cell.fill = win_fill
                elif value < 0:
                    cell.fill = loss_fill
                cell.number_format = "0.00"

        for idx, width in enumerate(
            [11, 14, 9, 11, 7, 10, 11, 10, 11, 9, 9, 10], start=1
        ):
            ws.column_dimensions[get_column_letter(idx)].width = width

    print(f"反包信号总数: {len(sig)} 笔")
    print(f"CSV : {OUT_CSV}")
    print(f"XLSX: {OUT_XLSX}")
    print()
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
