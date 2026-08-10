# -*- coding: utf-8 -*-
"""导出全年阴线炸板明细表(Excel + CSV)。

输入: results/zhaban_backtest_results_full.csv
输出: results/阴线炸板明细表_2025-04-21_2026-04-28.xlsx / .csv
"""

import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "results", "zhaban_backtest_results_full.csv")
OUT_XLSX = os.path.join(ROOT, "results", "阴线炸板明细表_2025-04-21_2026-04-28.xlsx")
OUT_CSV = os.path.join(ROOT, "results", "阴线炸板明细表_2025-04-21_2026-04-28.csv")


def main() -> None:
    df = pd.read_csv(SRC, encoding="utf-8-sig")
    df["time"] = pd.to_datetime(df["time"])

    # 全年阴线炸板: 炸板日收阴线(close < open), 其余筛选口径沿用回测脚本。
    yin = df[df["close"] < df["open"]].copy()
    yin = yin.sort_values(["time", "code"]).reset_index(drop=True)

    out = pd.DataFrame(
        {
            "日期": yin["time"].dt.strftime("%Y-%m-%d"),
            "代码": yin["code"],
            "炸板日收盘": yin["close"],
            "阴线": 1,
            "成交额(亿)": yin["成交额_亿"].round(2),
            "竞价涨幅(%)": yin["auction_pct"].round(2),
            "T-1涨停": yin["prev_limit_up"],
            "次日买入价": yin["next_open"],
            "次日卖出价": yin["next_close"],
            "盈亏(点)": (yin["next_close"] - yin["next_open"]).round(2),
            "盈亏(%)": yin["次日收益率_pct"].round(2),
        }
    )

    out.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="明细")
        ws = writer.sheets["明细"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="D9E2F3")
        win_fill = PatternFill("solid", fgColor="C6EFCE")
        loss_fill = PatternFill("solid", fgColor="FFC7CE")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
            for cell in row:
                try:
                    value = float(cell.value)
                except (TypeError, ValueError):
                    continue
                if value > 0:
                    cell.fill = win_fill
                elif value < 0:
                    cell.fill = loss_fill
                cell.number_format = "0.00"

        for idx, width in enumerate([11, 14, 10, 6, 10, 11, 8, 10, 10, 9, 9], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    print(f"阴线炸板明细: {len(out)} 笔")
    print(f"CSV : {OUT_CSV}")
    print(f"XLSX: {OUT_XLSX}")


if __name__ == "__main__":
    main()
