# -*- coding: utf-8 -*-
"""反包策略因子分组分析: 位置 / 成交额 / 连板数。

数据源: baostock 日线(2026-04-01 ~ 2026-07-31, 4月数据用于计算20日窗口和连板数)
信号: T-1 涨停, T 日未涨停; 买入 T+1 开盘, 卖出 T+2 收盘
统计: 笔数 / 胜率 / 平均收益 / 平均盈利 / 平均亏损 / 盈亏比
输出: results/反包因子分组_2026-05-01_2026-07-31.xlsx / .csv
"""

import io
import os
import sys
from decimal import Decimal, ROUND_HALF_UP

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BS_FILES = [
    os.path.join(ROOT, "data", "全市场A股_2026-04-01_2026-06-30_baostock.csv"),
    os.path.join(ROOT, "data", "全市场A股_2026-07-01_2026-07-31_baostock.csv"),
]
OUT_XLSX = os.path.join(ROOT, "results", "反包因子分组_2026-05-01_2026-07-31.xlsx")
OUT_CSV = os.path.join(ROOT, "results", "反包因子分组_2026-05-01_2026-07-31.csv")

POSITION_BINS = [-np.inf, 1 / 3, 2 / 3, np.inf]
POSITION_LABELS = ["低位(<=0.33)", "中位(0.33~0.66)", "高位(>0.66)"]
AMOUNT_BINS = [-np.inf, 1, 5, 10, 30, np.inf]
AMOUNT_LABELS = ["<1亿", "1-5亿", "5-10亿", "10-30亿", ">30亿"]


def _round_price(value: float, pct: float) -> float:
    """按交易所四舍五入规则计算涨停价。"""
    return float(
        (Decimal(str(value)) * Decimal(str(1 + pct))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    )


def _limit_pct(code: str, is_st: int) -> float:
    """按板块和 ST 状态返回涨跌幅限制。"""
    if code.startswith("bj."):
        return 0.30
    if code.startswith("sz.3") or code.startswith("sh.688"):
        return 0.20
    if is_st == 1:
        return 0.05
    return 0.10


def load_baostock(path: str) -> pd.DataFrame:
    """加载 baostock 日线并计算涨停标记、连板数、20日位置等因子。"""
    df = pd.read_csv(path, encoding="utf-8-sig", dtype={"code": str})
    for col in ["open", "high", "low", "close", "preclose", "amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["open", "close", "high", "low", "preclose", "amount"])
    df["成交额_亿"] = df["amount"] / 1e8
    df["high_limit"] = [
        _round_price(row.preclose, _limit_pct(row.code, int(row.isST)))
        for row in df.itertuples(index=False)
    ]
    df["is_limit_up"] = (df["close"] >= df["high_limit"] * 0.99).astype(int)
    return df


def calc_consecutive_limits(df: pd.DataFrame) -> pd.DataFrame:
    """计算每只股票连续涨停数(以当日为终点), 再取 T-1 的连板数。"""
    # 每次非涨停都把连板序列切断, 组内累计 is_limit_up 即连续涨停数
    group_key = (
        (df["is_limit_up"] == 0)
        .groupby(df["code"], sort=False)
        .cumsum()
    )
    df["连板数_当日"] = (
        df.groupby([df["code"], group_key], sort=False)["is_limit_up"].cumsum()
    )
    df["连板数_T1"] = df.groupby("code", sort=False)["连板数_当日"].shift(1)
    return df


def group_stats(df: pd.DataFrame, ret_col: str = "盈亏%") -> pd.Series:
    """计算一组交易的胜率、平均收益、平均盈利、平均亏损和盈亏比。"""
    ret = pd.to_numeric(df[ret_col], errors="coerce").dropna()
    n = len(ret)
    if n == 0:
        return pd.Series(
            {"笔数": 0, "胜率%": np.nan, "平均收益%": np.nan,
             "平均盈利%": np.nan, "平均亏损%": np.nan, "盈亏比": np.nan}
        )
    wins = ret[ret > 0]
    losses = ret[ret <= 0]
    avg_win = wins.mean() if len(wins) else np.nan
    avg_loss = losses.mean() if len(losses) else np.nan
    pl_ratio = abs(avg_win / avg_loss) if avg_loss else np.nan
    return pd.Series(
        {"笔数": n, "胜率%": round((ret > 0).mean() * 100, 1),
         "平均收益%": round(ret.mean(), 2),
         "平均盈利%": round(avg_win, 2) if avg_win == avg_win else np.nan,
         "平均亏损%": round(avg_loss, 2) if avg_loss == avg_loss else np.nan,
         "盈亏比": round(pl_ratio, 2) if pl_ratio == pl_ratio else np.nan}
    )


def summarize(df: pd.DataFrame, factor_col: str) -> pd.DataFrame:
    """按因子列分组汇总。"""
    out = (
        df.groupby(factor_col, observed=True)
        .apply(lambda x: group_stats(x), include_groups=False)
        .reset_index()
    )
    return out


def main() -> None:
    """合并数据、生成信号、按三个因子分组统计并导出。"""
    print("[运行] 合并 baostock 日线数据 ...")
    parts = [load_baostock(p) for p in BS_FILES if os.path.isfile(p)]
    if not parts:
        print("[错误] 未找到 baostock 数据文件")
        sys.exit(1)
    df = pd.concat(parts, ignore_index=True)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values(["code", "date"]).reset_index(drop=True)
    print(f"[运行] 原始日线 {len(df):,} 行, 股票 {df['code'].nunique()} 只")

    print("[运行] 计算连续涨停数与 20 日位置 ...")
    df = calc_consecutive_limits(df)
    grp = df.groupby("code", sort=False)
    df["low20"] = grp["low"].transform(lambda s: s.rolling(20, min_periods=20).min())
    df["high20"] = grp["high"].transform(lambda s: s.rolling(20, min_periods=20).max())
    df["位置"] = (df["close"] - df["low20"]) / (df["high20"] - df["low20"])
    df["prev_limit_up"] = grp["is_limit_up"].shift(1).fillna(0).astype(int)
    df["next_open"] = grp["open"].shift(-1)
    df["next2_close"] = grp["close"].shift(-2)

    print("[运行] 提取反包信号 (T-1涨停、T未涨停) ...")
    mask = (df["prev_limit_up"] == 1) & (df["is_limit_up"] == 0)
    sig = df[mask].dropna(subset=["next_open", "next2_close", "连板数_T1"]).copy()
    sig = sig[sig["date"] >= "2026-05-01"].copy()
    sig["盈亏%"] = (sig["next2_close"] / sig["next_open"] - 1) * 100
    sig["盈亏点"] = sig["next2_close"] - sig["next_open"]
    sig["连板档位"] = pd.cut(
        sig["连板数_T1"], bins=[-np.inf, 1, 2, np.inf],
        labels=["1板", "2板", "3板及以上"],
    )
    sig["位置档位"] = pd.cut(
        sig["位置"], bins=POSITION_BINS, labels=POSITION_LABELS, right=True
    )
    sig["成交额档位"] = pd.cut(
        sig["成交额_亿"], bins=AMOUNT_BINS, labels=AMOUNT_LABELS, right=False
    )
    print(f"[运行] 有效信号 {len(sig)} 笔")

    detail = pd.DataFrame(
        {
            "日期": sig["date"].dt.strftime("%Y-%m-%d"),
            "代码": sig["code"],
            "T日收盘": sig["close"].round(2),
            "连板数": sig["连板数_T1"].astype(int),
            "位置": sig["位置"].round(2),
            "成交额(亿)": sig["成交额_亿"].round(2),
            "次日买入价": sig["next_open"].round(2),
            "第三日卖出价": sig["next2_close"].round(2),
            "盈亏(点)": sig["盈亏点"].round(2),
            "盈亏(%)": sig["盈亏%"].round(2),
        }
    )
    detail = detail.sort_values(["日期", "代码"]).reset_index(drop=True)

    print("[运行] 分组统计 ...")
    overall = group_stats(sig).to_frame("全样本").T
    pos_sum = summarize(sig.assign(位置档位=sig["位置档位"]), "位置档位")
    amt_sum = summarize(sig.assign(成交额档位=sig["成交额档位"]), "成交额档位")
    streak_sum = summarize(sig.assign(连板档位=sig["连板档位"]), "连板档位")

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        detail.to_excel(writer, index=False, sheet_name="明细")
        overall.to_excel(writer, index=False, sheet_name="全样本")
        pos_sum.to_excel(writer, index=False, sheet_name="位置分组")
        amt_sum.to_excel(writer, index=False, sheet_name="成交额分组")
        streak_sum.to_excel(writer, index=False, sheet_name="连板数分组")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx in range(1, ws.max_column + 1):
                widths = [
                    len(str(cell.value)) * 2 + 2
                    for cell in ws[col_idx] if cell.value is not None
                ]
                width = max(widths) if widths else 10
                ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 18)

    detail.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print(f"[完成] CSV : {OUT_CSV}")
    print(f"[完成] XLSX: {OUT_XLSX}")
    print()

    def show(title: str, table: pd.DataFrame) -> None:
        print(f"===== {title} =====")
        print(table.to_string(index=False))
        print()

    show("全样本", overall)
    show("按位置分组", pos_sum)
    show("按成交额分组", amt_sum)
    show("按连板数分组", streak_sum)


if __name__ == "__main__":
    main()
